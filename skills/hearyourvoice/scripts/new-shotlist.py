#!/usr/bin/env python3
"""new-shotlist.py — generate a blank 5-sheet shotlist.xlsx for a HearYourVOICE project.

Mirrors the chado example: Shot List, Capture Plan, Episode Coverage,
YouTube CC Sources, Source Strategy. See references/shotlist-format.md.

Usage:
  python3 new-shotlist.py --slug my-topic --episodes 10
  python3 new-shotlist.py --slug my-topic --out src/my-topic/shotlist.xlsx

Needs openpyxl:  pip install openpyxl
"""
import argparse
import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

HEAD_FILL = PatternFill("solid", fgColor="2A3B3B")
HEAD_FONT = Font(bold=True, color="FFFFFF")
SECTION_FONT = Font(bold=True)


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"


def widths(ws, ws_widths):
    for i, w in enumerate(ws_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build(slug, episodes, out):
    wb = Workbook()

    # 1. Shot List
    ws = wb.active
    ws.title = "Shot List"
    head = ["ID", "หมวด", "ช็อต (Shot)", "ตอน (EP)", "ลำดับ",
            "วิธีถ่าย / โน้ต", "สถานะ", "ที่หมาย / วันถ่าย",
            "แหล่ง Footage แนะนำ", "ลิงก์/Keyword", "License/หมายเหตุ"]
    ws.append(head)
    ws.append(["UW-01", "Underwater", "<describe the shot>", "EP1", "P1",
               "<how to capture / direction>", "ยังไม่ถ่าย", "",
               "self-shot / CC / stock / generative / graphics", "<url or keyword>", "<license note>"])
    ws.append(["GFX-01", "Graphics", "<diagram / typography beat>", "EP1", "P2",
               "Remotion-native, post", "—", "", "graphics", "", "made in post"])
    style_header(ws, len(head))
    widths(ws, [10, 14, 38, 12, 8, 30, 12, 16, 26, 30, 28])

    # 2. Capture Plan
    ws = wb.create_sheet("Capture Plan")
    ws.append(["แผนเก็บภาพ (จัดตามชุดการถ่าย — ไว้ batch / หา insert ล่วงหน้า)"])
    ws["A1"].font = HEAD_FONT
    ws["A1"].fill = HEAD_FILL
    rows = [
        [""],
        ["เซสชัน A — สตูดิโอ / Macro", ""],
        ["เมื่อไร / เงื่อนไข:", "<anytime, controlled light>"],
        ["ช็อตในชุด:", "<shot IDs, e.g. LRM-01..LRM-18>"],
        [""],
        ["เซสชัน B — ผิวน้ำ / golden hour", ""],
        ["เมื่อไร / เงื่อนไข:", "<golden hour, on location>"],
        ["ช็อตในชุด:", "<shot IDs>"],
        [""],
        ["เซสชัน C — <name>", ""],
        ["เมื่อไร / เงื่อนไข:", ""],
        ["ช็อตในชุด:", ""],
    ]
    for r in rows:
        ws.append(r)
    for r in range(2, ws.max_row + 1):
        a = ws.cell(row=r, column=1)
        if isinstance(a.value, str) and a.value.startswith("เซสชัน"):
            a.font = SECTION_FONT
    widths(ws, [28, 60])

    # 3. Episode Coverage
    ws = wb.create_sheet("Episode Coverage")
    head = ["ตอน", "ชื่อตอน", "ช็อตที่ต้องใช้", "จำนวน"]
    ws.append(head)
    for i in range(1, episodes + 1):
        ws.append([f"EP{i}", "<episode title>", "<shot IDs, comma-separated>", ""])
    style_header(ws, len(head))
    widths(ws, [8, 30, 50, 8])

    # 4. YouTube CC Sources
    ws = wb.create_sheet("YouTube CC Sources")
    head = ["Priority", "Use case", "Title", "Channel", "Length", "URL", "Matched EP/Shot", "License note"]
    ws.append(head)
    ws.append(["A", "<use case>", "<video title>", "<channel>", "0:00", "<url>", "EP1 / UW-01",
               "Verify CC BY via yt-dlp before use"])
    style_header(ws, len(head))
    widths(ws, [9, 26, 36, 20, 8, 40, 18, 34])

    # 5. Source Strategy
    ws = wb.create_sheet("Source Strategy")
    head = ["ประเภทภาพ", "แหล่งที่เหมาะ", "หมายเหตุ"]
    ws.append(head)
    for row in [
        ["Hero behavior / action", "YouTube CC + creator permission + field shoot", "CC ok w/ credit; DM if prominent"],
        ["Close-up / detail", "stock + controlled tank/studio", "stock = buy license"],
        ["Macro / product", "self-shot studio", "cheapest, best control"],
        ["Action / process", "self-shot pond/tank/location", "no live subject needed to convey action"],
        ["Habitat / mood", "self-shot golden hour + drone/gimbal", "keeps series tone consistent"],
        ["Graphics / data", "Remotion post", "for shots too hard to film"],
    ]:
        ws.append(row)
    style_header(ws, len(head))
    widths(ws, [28, 40, 40])

    os.makedirs(os.path.dirname(os.path.abspath(out)) or ".", exist_ok=True)
    wb.save(out)
    print(f"Wrote {out}  (sheets: {', '.join(wb.sheetnames)}; {episodes} episodes)")


def main():
    ap = argparse.ArgumentParser(description="Generate a blank HearYourVOICE shotlist.xlsx")
    ap.add_argument("--slug", required=True)
    ap.add_argument("--episodes", type=int, default=1)
    ap.add_argument("--out", default=None)
    a = ap.parse_args()
    out = a.out or os.path.join("src", a.slug, "shotlist.xlsx")
    build(a.slug, max(1, a.episodes), out)


if __name__ == "__main__":
    main()
